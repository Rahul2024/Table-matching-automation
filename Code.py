import pandas as pd
import numpy as np
from fuzzywuzzy import process, fuzz
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


# === Config ===
file_path = "/Users/rahulraj/Downloads/MMS05_G_N_1440_AO.xlsx"
key_columns = ['Purchasing Document'    , 'Purchasing Document Item' ,     'Delivery No','Delivery Schedule Line']
output_file = "/Users/rahulraj/desktop/MMS05_G_N_1440_AO_final.xlsx"



# === FUZZY MATCHING SETTINGS ===
MIN_FUZZY_SCORE = 85  # Increased from 92 for better matching
EXACT_MATCH_BONUS = 10  # Bonus for exact substring matches

# === Load Sheets ===
sheet1 = pd.read_excel(file_path, sheet_name='Sheet1')
sheet2 = pd.read_excel(file_path, sheet_name='Sheet2')

# ---Column Cleaning ---
def clean_columns(df):
    """Clean column names more thoroughly."""
    cleaned_cols = []
    for col in df.columns:
        clean_col = str(col).strip().replace('\u00A0', ' ').replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        while '  ' in clean_col:
            clean_col = clean_col.replace('  ', ' ')
        cleaned_cols.append(clean_col.strip())
    return cleaned_cols

sheet1.columns = clean_columns(sheet1)
sheet2.columns = clean_columns(sheet2)

# === IMPROVED KEY COLUMN VALIDATION ===
def find_key_column_improved(df_columns, target_key):
    """Find key column using exact match first, then comprehensive fuzzy matching."""
    # First try exact match
    if target_key in df_columns:
        return target_key, 100
    
    if not df_columns:
        return None, 0
    
    # Try multiple fuzzy matching strategies
    best_match = None
    best_score = 0
    
    for col in df_columns:
        # Strategy 1: Standard ratio
        score1 = fuzz.ratio(target_key.lower(), col.lower())
        
        # Strategy 2: Partial ratio (good for substring matches)
        score2 = fuzz.partial_ratio(target_key.lower(), col.lower())
        
        # Strategy 3: Token sort ratio (good for word order differences)
        score3 = fuzz.token_sort_ratio(target_key.lower(), col.lower())
        
        # Strategy 4: Token set ratio (good for different word sets)
        score4 = fuzz.token_set_ratio(target_key.lower(), col.lower())
        
        # Take the maximum score from all strategies
        max_score = max(score1 ,  score2 , score3 , score4)
        
        # Bonus for exact substring matches
        if target_key.lower() in col.lower() or col.lower() in target_key.lower():
            max_score = min(100, max_score + EXACT_MATCH_BONUS)
        
        if max_score > best_score:
            best_score = max_score
            best_match = col
    
    return best_match, best_score

# === DUPLICATE COLUMN CHECK ===
dupe_cols = sheet1.columns[sheet1.columns.duplicated()].tolist()
if dupe_cols:
    print("❌ Duplicate columns in Sheet1:", dupe_cols)
    raise ValueError('❌ Duplicate columns in Sheet1 ❌')

dupe_2cols = sheet2.columns[sheet2.columns.duplicated()].tolist()
if dupe_2cols:
    print("❌ Duplicate columns in Sheet2:", dupe_2cols)
    raise ValueError('❌ Duplicate columns in Sheet2 ❌')

# === KEY COLUMN MATCHING ===
actual_key_columns = []
for key in key_columns:
    found_key1, score1 = find_key_column_improved(sheet1.columns.tolist(), key)
    found_key2, score2 = find_key_column_improved(sheet2.columns.tolist(), key)
    
    if found_key1 and found_key2 and score1 >= MIN_FUZZY_SCORE and score2 >= MIN_FUZZY_SCORE:
        actual_key_columns.append(key)
        if found_key1 != key: 
            print(f"🔄 Renaming '{found_key1}' to '{key}' in Sheet1 (score: {score1})")
            sheet1.rename(columns={found_key1: key}, inplace=True)
        if found_key2 != key: 
            print(f"🔄 Renaming '{found_key2}' to '{key}' in Sheet2 (score: {score2})")
            sheet2.rename(columns={found_key2: key}, inplace=True)
    else:
        print(f"❌ Error: Key column '{key}' not found!")
        print(f"   Sheet1 best match: '{found_key1}' (score: {score1})")
        print(f"   Sheet2 best match: '{found_key2}' (score: {score2})")
        raise Exception(f"Missing key column: {key}")

key_columns = actual_key_columns

# === Store Original Key Values Before Processing ===
original_keys_sheet1 = {key: sheet1[key].copy() for key in key_columns}
original_keys_sheet2 = {key: sheet2[key].copy() for key in key_columns}

# === Fixed Normalize Key Columns ===
def normalize_key_value(val):
    """Normalize key values to clean string format without .0 suffix."""
    if pd.isna(val): 
        return str(val)
    
    str_val = str(val).strip()
    
    if str_val.endswith('.0'):
        str_val = str_val[:-2]
    
    if str_val.replace('.', '').replace('-', '').isdigit():
        try:
            return str(int(float(str_val)))
        except (ValueError, OverflowError):
            return str_val
    
    return str_val

for col in key_columns:
    sheet1[f'{col}_normalized'] = sheet1[col].apply(normalize_key_value)
    sheet2[f'{col}_normalized'] = sheet2[col].apply(normalize_key_value)
normalized_key_cols = [f'{col}_normalized' for col in key_columns]

# === Create Unique Composite Keys ===
sheet1['__key__'] = sheet1[normalized_key_cols].agg('|'.join, axis=1)
sheet2['__key__'] = sheet2[normalized_key_cols].agg('|'.join, axis=1)

if sheet1['__key__'].duplicated().any():
    sheet1['__key__'] += '_row' + (sheet1.index + 1).astype(str)
if sheet2['__key__'].duplicated().any():
    sheet2['__key__'] += '_row' + (sheet2.index + 1).astype(str)

# === NEW: Identify Common and Extra Rows ===
sheet1_keys = set(sheet1['__key__'])
sheet2_keys = set(sheet2['__key__'])

# Find common keys and extra keys
common_keys = sheet1_keys & sheet2_keys
extra_in_sheet1 = sheet1_keys - sheet2_keys  # In Sheet1 but not in Sheet2
extra_in_sheet2 = sheet2_keys - sheet1_keys  # In Sheet2 but not in Sheet1

print(f"📊 Key Analysis:")
print(f"  - Total keys in Sheet1: {len(sheet1_keys)}")
print(f"  - Total keys in Sheet2: {len(sheet2_keys)}")
print(f"  - Common keys (for comparison): {len(common_keys)}")
print(f"  - Extra in Sheet1 only: {len(extra_in_sheet1)}")
print(f"  - Extra in Sheet2 only: {len(extra_in_sheet2)}")

# === Clean up normalized columns ===
for col in normalized_key_cols:
    if col in sheet1.columns:
        sheet1.drop(columns=[col], inplace=True)
    if col in sheet2.columns:
        sheet2.drop(columns=[col], inplace=True)

sheet1.set_index('__key__', inplace=True)
sheet2.set_index('__key__', inplace=True)

# === NEW: Filter to Common Keys Only for Comparison ===
sheet1_common = sheet1.loc[list(common_keys)]
sheet2_common = sheet2.loc[list(common_keys)]

# === IMPROVED FUZZY MATCH COLUMNS (UPDATED SECTION) ===
# Only include actual data columns, exclude key columns and any remaining normalized columns
sheet1_cols = [c for c in sheet1_common.columns if c not in key_columns and not c.endswith('_normalized')]
sheet2_cols = [c for c in sheet2_common.columns if c not in key_columns and not c.endswith('_normalized')]

print(f"\n🔍 Starting fuzzy column matching...")
print(f"   Source columns: {len(sheet1_cols)}")
print(f"   Target columns: {len(sheet2_cols)}")

matched_cols, unmatched_cols = {}, []
for col1 in sheet1_cols:
    match, score = process.extractOne(col1, sheet2_cols)
    if score >= 92:  # changed the factor by 93 from 90 for better accuracy
        matched_cols[col1] = match
        print(f"   ✅ Matched: '{col1}' → '{match}' (score: {score})")
    else:
        unmatched_cols.append(col1)
        print(f"   ❌ No match: '{col1}' (best: '{match}', score: {score})")

# Find unmatched columns in sheet2
matched_sheet2_cols = set(matched_cols.values())
unmatched_sheet2 = [col for col in sheet2_cols if col not in matched_sheet2_cols]

print(f"\n📊 Matching Results:")
print(f"   ✅ Successfully matched: {len(matched_cols)} pairs")
print(f"   ❌ Unmatched in Sheet1: {len(unmatched_cols)}")
print(f"   ❌ Unmatched in Sheet2: {len(unmatched_sheet2)}")

if unmatched_cols:
    print(f"   📝 Unmatched Sheet1 columns: {unmatched_cols}")
if unmatched_sheet2:
    print(f"   📝 Unmatched Sheet2 columns: {unmatched_sheet2}")

# === Build Side-by-side Sheet (Common Keys Only) ===
sheet1_comparison_result = pd.DataFrame(index=sheet1_common.index)
for col1 in sheet1_cols:
    sheet1_comparison_result[col1] = sheet1_common[col1]
    if col1 in matched_cols:
        col2 = matched_cols[col1]
        sheet1_comparison_result[f"{col2} (target)"] = sheet2_common[col2]
    else:
        sheet1_comparison_result[f"Missing (target)"] = "Missing in Target"

# Insert composite key
sheet1_comparison_result.insert(0, '__key__', sheet1_comparison_result.index)

# Create mapping for original key values (common keys only)
key_to_original_sheet1 = {}
key_to_original_sheet2 = {}

# Get original indices for common keys
original_sheet1_indices = []
original_sheet2_indices = []

for comp_key in common_keys:
    # Find original index in sheet1
    orig_idx1 = sheet1.index.get_loc(comp_key)
    original_sheet1_indices.append(orig_idx1)
    key_to_original_sheet1[comp_key] = {key: original_keys_sheet1[key].iloc[orig_idx1] for key in key_columns}
    
    # Find original index in sheet2  
    orig_idx2 = sheet2.index.get_loc(comp_key)
    original_sheet2_indices.append(orig_idx2)
    key_to_original_sheet2[comp_key] = {key: original_keys_sheet2[key].iloc[orig_idx2] for key in key_columns}

# Insert original key columns
for idx, key in enumerate(reversed(key_columns), 1):
    original_values = [key_to_original_sheet1[comp_key][key] for comp_key in sheet1_comparison_result.index]
    sheet1_comparison_result.insert(idx, key, original_values)

def are_values_equal_enhanced(v1, v2):
    """Enhanced comparison with more robust empty value handling."""
    
    def normalize_value(val):
        """Normalize a value to a standard form for comparison."""
        if pd.isna(val) or val is None:
            return None
        
        # Handle numeric values
        if isinstance(val, (int, float)):
            if val == 0 or val == 0.0 or val== 0.00 :
                return 0
            return val
        
        # Handle string values
        if isinstance(val, str):
            cleaned = val.strip().lower()
            
            # Check for various empty representations
            if cleaned in ['', '0', '0.00','0.0', '#', 'nan', 'null', 'none', 'not assigned', '-']:
                return None
            
            # Try to convert to number if it looks like one
            try:
                num_val = float(cleaned)
                if num_val == 0.0:
                    return 0
                return num_val
            except (ValueError, TypeError):
                return cleaned
        
        # Try to convert other types to string and normalize
        str_val = str(val).strip().lower()
        if str_val in ['', '0', '0.00'  , '0.0', 'nan', 'null', 'none', 'not assigned', '-']:
            return None
        
        try:
            num_val = float(str_val)
            if num_val == 0.0:
                return 0
            return num_val
        except (ValueError, TypeError):
            return str_val
    
    # Normalize both values
    norm_v1 = normalize_value(v1)
    norm_v2 = normalize_value(v2)
    
    # Compare normalized values
    if norm_v1 is None and norm_v2 is None:
        return True
    if norm_v1 is None or norm_v2 is None:
        return False
    if norm_v1 == 0 and norm_v2 == 0:
        return True
    
    return norm_v1 == norm_v2

# === Build Comparison (Common Keys Only) ===
comparison = pd.DataFrame(index=sheet1_common.index)
for col1, col2 in matched_cols.items():
    val1 = sheet1_common[col1]
    val2 = sheet2_common[col2]
    comparison[col1] = [are_values_equal_enhanced(val1.iloc[i], val2.iloc[i]) for i in range(len(val1))]

for col1 in unmatched_cols:
    comparison[col1] = "Missing in Sheet2"

# === Overall result Sheet ===
column_status = []
for col2 in unmatched_sheet2: 
    column_status.append({'Column': col2, 'Status': 'Missing in Source', 'KPI': 'FAIL'})
for col1 in unmatched_cols: 
    column_status.append({'Column': col1, 'Status': 'Missing in Target', 'KPI': 'FAIL'})
for col in matched_cols:
    if comparison[col].dtype == bool:
        if comparison[col].all():
            column_status.append({'Column': col, 'Status': 'All values match', 'KPI': 'PASS'})
        else:
            mismatch_count = (~comparison[col]).sum()
            column_status.append({'Column': col, 'Status': f'Mismatches: {mismatch_count}/{len(comparison[col])}', 'KPI': 'FAIL'})
column_status_df = pd.DataFrame(column_status)

# === Column Mapping Sheet ===
column_comparison_data = []
for col1, col2 in matched_cols.items(): 
    # Get the actual score from the matching process
    _, score = process.extractOne(col1, sheet2_cols)
    column_comparison_data.append({'Source_Column': col1, 'Target_Column': col2, 'Match_Status': 'Matched', 'Fuzzy_Score': score})

for col1 in unmatched_cols: 
    column_comparison_data.append({'Source_Column': col1, 'Target_Column': 'Not Found', 'Match_Status': 'Missing in Target', 'Fuzzy_Score': 'N/A'})
for col2 in unmatched_sheet2: 
    column_comparison_data.append({'Source_Column': 'Not Found', 'Target_Column': col2, 'Match_Status': 'Missing in Source', 'Fuzzy_Score': 'N/A'})
column_comparison_df = pd.DataFrame(column_comparison_data)

# === Mismatch Details (Common Keys Only) ===
mismatch_rows_mask = (comparison == False).any(axis=1)

# NEW: Use Overall Result sheet to identify failed columns
failed_columns = column_status_df[column_status_df['KPI'] == 'FAIL']['Column'].tolist()

# Build final mismatch columns list - only include failed columns
final_mismatch_cols = ['__key__'] + key_columns

# Add only failed columns (both source and target versions)
for col1 in failed_columns:
    if col1 in sheet1_comparison_result.columns:
        final_mismatch_cols.append(col1)
        # Add corresponding target column if it exists
        if col1 in matched_cols:
            col2 = matched_cols[col1]
            target_col_name = f"{col2} (target)"
            if target_col_name in sheet1_comparison_result.columns:
                final_mismatch_cols.append(target_col_name)

# Filter to only available columns
available_cols = [col for col in final_mismatch_cols if col in sheet1_comparison_result.columns]
mismatch_df = sheet1_comparison_result.loc[mismatch_rows_mask, available_cols].copy()

# === NEW: Create Extra Rows Sheet ===
extra_rows_data = []

# Add extra rows from Sheet1 (Yellow)
if extra_in_sheet1:
    for key in extra_in_sheet1:
        row_data = {'__key__': key, 'Source': 'Sheet1 Only'}
        
        # Add original key values
        orig_idx = sheet1.index.get_loc(key)
        for col in key_columns:
            row_data[col] = original_keys_sheet1[col].iloc[orig_idx]
        
        # Add all other columns from sheet1
        for col in sheet1.columns:
            if col not in key_columns:
                row_data[f'{col} (Sheet1)'] = sheet1.loc[key, col]
        
        extra_rows_data.append(row_data)

# Add extra rows from Sheet2 (Blue)
if extra_in_sheet2:
    for key in extra_in_sheet2:
        row_data = {'__key__': key, 'Source': 'Sheet2 Only'}
        
        # Add original key values
        orig_idx = sheet2.index.get_loc(key)
        for col in key_columns:
            row_data[col] = original_keys_sheet2[col].iloc[orig_idx]
        
        # Add all other columns from sheet2
        for col in sheet2.columns:
            if col not in key_columns:
                row_data[f'{col} (Sheet2)'] = sheet2.loc[key, col]
        
        extra_rows_data.append(row_data)

# Create DataFrame for extra rows
if extra_rows_data:
    extra_rows_df = pd.DataFrame(extra_rows_data)
    # Reorder columns to have key columns first
    cols_order = ['__key__', 'Source'] + key_columns + [col for col in extra_rows_df.columns if col not in (['__key__', 'Source'] + key_columns)]
    extra_rows_df = extra_rows_df.reindex(columns=cols_order)
else:
    extra_rows_df = pd.DataFrame()

# === Write to Excel ===
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    sheet1.reset_index().to_excel(writer, sheet_name="Source Data", index=False)
    sheet2.reset_index().to_excel(writer, sheet_name="Target Data", index=False)
    comparison.reset_index().to_excel(writer, sheet_name="Row Comparison", index=False)
    column_status_df.to_excel(writer, sheet_name="Overall Result", index=False)
    column_comparison_df.to_excel(writer, sheet_name="Column Mapping", index=False)
    sheet1_comparison_result.reset_index(drop=True).to_excel(writer, sheet_name="Side by Side Result", index=False)
    mismatch_df.reset_index(drop=True).to_excel(writer, sheet_name="Mismatch Details", index=False)
    extra_rows_df.to_excel(writer, sheet_name="Extra Rows Analysis", index=False)  # NEW SHEET

# === Apply color formatting ===
wb = load_workbook(output_file)
orange_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Sheet1 extra rows
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")    # Sheet2 extra rows

def auto_fit_columns_simple(worksheet):
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 4

sheet_names = wb.sheetnames
for sheet_name in sheet_names:
    auto_fit_columns_simple(wb[sheet_name])

# Color Side by Side Result
ws_sbs = wb["Side by Side Result"]
for col_idx in range(1, ws_sbs.max_column + 1):
    header = ws_sbs.cell(row=1, column=col_idx).value
    if not header: continue
    fill = None
    if header in key_columns or header == '__key__': fill = orange_fill
    elif header.endswith("(target)"): fill = grey_fill
    if fill:
        for row_num in range(1, ws_sbs.max_row + 1):
            ws_sbs[f"{get_column_letter(col_idx)}{row_num}"].fill = fill

# Color Overall Result
ws_or = wb["Overall Result"]
for row in ws_or.iter_rows(min_row=2):
    kpi = row[2].value
    fill = green_fill if kpi == "PASS" else red_fill if kpi == "FAIL" else None
    if fill:
        for cell in row: cell.fill = fill

# Color Column Mapping
ws_cm = wb["Column Mapping"]
for row in ws_cm.iter_rows(min_row=2):
    status = row[2].value
    fill = green_fill if status == "Matched" else red_fill if "Missing" in status else None
    if fill:
        for cell in row: cell.fill = fill

# Color Mismatch Details Sheet
if "Mismatch Details" in wb.sheetnames and len(mismatch_df) > 0:
    ws_md = wb["Mismatch Details"]
    headers = [cell.value for cell in ws_md[1]]
    
    mismatch_comparison_df = comparison[mismatch_rows_mask].copy()

    for r_idx, row in enumerate(ws_md.iter_rows(min_row=2), start=2):
        row_key = row[0].value
        for c_idx, cell in enumerate(row, start=1):
            header = headers[c_idx - 1]
            original_col_name = header.replace(" (target)", "")

            if original_col_name in mismatch_comparison_df.columns:
                is_match = mismatch_comparison_df.loc[row_key, original_col_name]
                if is_match == False:
                    cell.fill = red_fill

# NEW: Color Extra Rows Analysis Sheet
if "Extra Rows Analysis" in wb.sheetnames and len(extra_rows_df) > 0:
    ws_era = wb["Extra Rows Analysis"]
    
    # Color header row (Source column in orange, key columns in orange)
    for col_idx in range(1, ws_era.max_column + 1):
        header = ws_era.cell(row=1, column=col_idx).value
        if header in key_columns or header == '__key__' or header == 'Source':
            ws_era.cell(row=1, column=col_idx).fill = orange_fill
    
    # Color data rows based on source
    for row_num in range(2, ws_era.max_row + 1):
        source_cell = ws_era.cell(row=row_num, column=2)  # 'Source' column
        if source_cell.value == "Sheet1 Only":
            fill = yellow_fill
        elif source_cell.value == "Sheet2 Only":
            fill = blue_fill
        else:
            continue
            
        # Apply color to entire row
        for col_num in range(1, ws_era.max_column + 1):
            ws_era.cell(row=row_num, column=col_num).fill = fill

wb.save(output_file)
print("Key columns:", key_columns)
print("Sample common keys:")
for i, key in enumerate(list(common_keys)[:5]):
    print(f"  {key}")
print(f"📊 Final Summary:")
print(f"  - Common rows compared: {len(common_keys)}")
print(f"  - Extra in Sheet1 (Yellow): {len(extra_in_sheet1)}")
print(f"  - Extra in Sheet2 (Blue): {len(extra_in_sheet2)}")
print(f"✨ Abracadabra! File magically appeared at '{output_file}' 🪄")
